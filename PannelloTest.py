import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import List
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import re
import xlsxwriter
import pandas as pd
import pyvisa
from tkinter import filedialog
import math
from pymodbus.client import ModbusTcpClient, ModbusSerialClient
from pymodbus.exceptions import ModbusIOException
import ast  # per convertire stringa in lista, se serve
import os
import openpyxl
import threading
import time
import csv
from datetime import datetime
from drivers.instruments import *
from drivers.decoders import decode_u16_auto
from drivers.test import *
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from collections import deque, defaultdict
import threading

# buffer realtime condiviso
rt_columns = []  # lista colonne (es. ["Inverter1_Voltage DC1 [V]", ...])
rt_time = deque(maxlen=2000)  # timestamp (uno per riga)
rt_data = defaultdict(lambda: deque(maxlen=2000))  # colname -> deque di valori
rt_lock = threading.Lock()

# riferimenti globali per gestire run successivi
current_shared_ins = None
logging_thread = None
test_thread = None
logging_running = False
logging_paused = False
current_report_ctx = {}  # {"template_path": "...", "serials": [...]} oppure vuoto
current_session_dir = None  # cartella della sessione corrente .\Data\SN_YYYYMMDD_HHMMSS


def _visa_is_present(addr: str, timeout_ms: int = 500) -> tuple[bool, str]:
    """Prova ad aprire la risorsa VISA 'addr'. True se apre, False se fallisce."""
    try:
        rm = pyvisa.ResourceManager()
        inst = rm.open_resource(addr)
        try:
            inst.timeout = timeout_ms
            # Tentativo soft di *IDN? (se non supportato non è un errore fatale)
            try:
                idn = inst.query("*IDN?").strip()
            except Exception:
                idn = ""
        finally:
            try: inst.close()
            except Exception: pass
        return True, idn
    except Exception as e:
        return False, str(e)


# ------------------------------------------------------------
# Stima durata test (per impostare il tempo del logging)
# ------------------------------------------------------------
def estimate_total_time(template_path: str, sn_for_test: str, template_folder: str = "./template") -> float:
    """
    Ritorna la durata stimata del test in secondi, per impostare il logging.
    Regole:
      - custom/altro: somma della colonna 'tempo'
      - 'curva MPPT' nel nome file: df_template['tempo'][0] * (Vmax/5)
    """
    try:
        df_template = pd.read_excel(template_path)
    except Exception:
        return 60.0  # fallback

    # Caso 'curva MPPT' (rilevato dal nome file come nel tuo codice)
    if "curva MPPT" in os.path.basename(template_path):
        # step (s): prima cella di 'tempo' o 1.0 se non presente
        step_s = 1.0
        if "tempo" in df_template.columns:
            try:
                step_s = float(str(df_template["tempo"].iloc[0]).replace(",", "."))
            except Exception:
                step_s = 1.0

        # ricava Vmax dal DB in base al SN (family/model_code)
        try:
            info = parse_sn(sn_for_test)
            family     = info["family"][0] if isinstance(info["family"], (list, tuple)) else info["family"]
            model_code = info["model_code"][0] if isinstance(info["model_code"], (list, tuple)) else info["model_code"]
            db_path = os.path.join("./database", f"{family}.xlsx")
            # leggi chiave come stringa e normalizza a 3 cifre
            df_inv = pd.read_excel(db_path, dtype={"Unnamed: 0": str}, keep_default_na=False)
            df_inv["Unnamed: 0"] = (
                df_inv["Unnamed: 0"].astype(str).str.strip()
                .str.replace(r"\.0$", "", regex=True).str.zfill(3)
            )
            key = str(model_code).zfill(3)
            vmax = float(df_inv.loc[df_inv["Unnamed: 0"] == key, "MAX V"].iloc[0])
        except Exception:
            vmax = 100.0  # fallback prudente

        # tua regola: tempo = step * (Vmax / 5)
        return float(step_s) * (float(vmax) / 5.0)

    # Caso 'custom' / altri template: somma dei tempi
    if "tempo" in df_template.columns:
        try:
            tempi = pd.to_numeric(df_template["tempo"].astype(str).str.replace(",", "."), errors="coerce").fillna(0.0)
            return float(tempi.sum())
        except Exception:
            return 60.0
    return 60.0


# stop sicuro del logger  rilascio COM/IP
def stop_logging_and_release():
    global logging_running, logging_thread, current_shared_ins
    try:
        logging_running = False
        t = logging_thread
        logging_thread = None
        if t and t.is_alive():
            try: t.join(timeout=3.0)
            except Exception as e: print(f"[WARN] join logging: {e}")
        if current_shared_ins:
            try: current_shared_ins.inv_disconnect_all()
            except Exception as e: print(f"[WARN] inv_disconnect_all: {e}")
        time.sleep(0.3)  # lascia Windows rilasciare la COM
    except Exception as e:
        print(f"[WARN] stop_logging_and_release: {e}")


def open_realtime_panel(colnames, default_col=None):
    # registra le colonne mostrate nel pannello
    global rt_columns
    rt_columns.clear(); rt_columns.extend(colnames)

    win = tk.Toplevel()
    win.title("Monitoraggio Realtime")
    win.minsize(900, 600)

    # TOP: selezione colonna
    top = tk.Frame(win)
    top.pack(fill="x", padx=10, pady=8)

    tk.Label(top, text="Grandezza:").pack(side="left")
    selected_col = tk.StringVar(value=default_col or (rt_columns[0] if rt_columns else ""))
    col_combo = ttk.Combobox(top, textvariable=selected_col, values=rt_columns, width=60, state="readonly")
    col_combo.pack(side="left", padx=8)

    # Pulsanti pause/resume/exit
    def _pause():
        global logging_paused
        logging_paused = True
    def _resume():
        global logging_paused
        logging_paused = False

    tk.Button(top, text="Pause",  command=_pause, bg='orange').pack(side="right", padx=4)
    tk.Button(top, text="Resume", command=_resume, bg='lightgreen').pack(side="right", padx=4)
    tk.Button(top, text="Exit",   command=win.destroy, bg='red').pack(side="right", padx=4)

    # MIDDLE: grafico matplotlib
    fig = Figure(figsize=(6,4), dpi=100)
    ax = fig.add_subplot(111)
    ax.set_xlabel("tempo")
    ax.set_ylabel("valore")
    line, = ax.plot([], [])  # iniziale vuota
    canvas = FigureCanvasTkAgg(fig, master=win)
    canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=6)

    # BOTTOM: tabella ultimi 10
    table_frame = tk.LabelFrame(win, text="Ultimi 10 campioni")
    table_frame.pack(fill="x", padx=10, pady=6)
    # tree = ttk.Treeview(table_frame, columns=("ts","val"), show="headings", height=10)
    # tree.heading("ts",  text="Timestamp")
    # tree.heading("val", text="Valore")
    # tree.column("ts",  width=200, anchor="w")
    # tree.column("val", width=150, anchor="e")

    # stile tabella
    style = ttk.Style(table_frame)
    style.configure("Log.Treeview", font=("Consolas", 10), rowheight=22)
    style.configure("Log.Treeview.Heading", font=("Arial", 10, "bold"))
    tree = ttk.Treeview(table_frame, columns=("ts","val"), show="headings", height=10, style="Log.Treeview")
    tree.heading("ts", text="Timestamp")
    tree.heading("val", text="Valore")
    tree.column("ts", width=160, anchor="w")
    tree.column("val", width=120, anchor="e")  # numeri a destra
    tree.tag_configure("odd",  background="#f6f7fb")
    tree.tag_configure("even", background="#ffffff")

    tree.pack(fill="x")

    def _update_view():
        col = selected_col.get()
        with rt_lock:
            # copia dati correnti (evita race)
            ts_list = list(rt_time)
            ys = list(rt_data.get(col, []))

        # converti x in indici (mostra ultimi N punti)
        if ts_list and ys:
            n = len(ys)
            start = max(0, n - 300)
            xs = list(range(start, n))  # asse x progressivo
            yv = ys[start:]  # include anche NaN → matplotlib spezza la linea
            line.set_data(xs, yv)
            ax.relim(); ax.autoscale_view()
            canvas.draw_idle()

            # tabella ultimi 10
            for item in tree.get_children():
                tree.delete(item)
            last10_vals = yv[-10:]
            # prendi gli ultimi timestamp, ma limita alla stessa cardinalità dei valori
            last10_ts = ts_list[-len(last10_vals):] if ts_list else []
            # padding sicuro (pad sempre definito)
            pad = max(0, len(last10_vals) - len(last10_ts))
            if pad:
                last10_ts = ([""] * pad) + last10_ts
            for idx, (ts, val) in enumerate(zip(last10_ts, last10_vals)):
                disp = "-" if (isinstance(val, float) and math.isnan(val)) else val
                tag = "even" if (idx % 2 == 0) else "odd"
                tree.insert("", "end", values=(ts, disp), tags=(tag,))
        win.after(500, _update_view)  # refresh ogni 0.5 s

    _update_view()

    def _on_change(*_):
        # reset della vista (facoltativo)
        pass
    col_combo.bind("<<ComboboxSelected>>", _on_change)


# --- helper da hex a int ---
def _to_int_reg(r):
    if isinstance(r, int): return r
    s = str(r).strip().lower()
    return int(s, 16) if s.startswith("0x") else int(s)


# --- lettura SN Inverter ---
def read_SN(proto, ip_tcp=None, porta_com=None, ip_hub=None,
            slave_id_rtu=None, slave_id_azzurro=None, max_retries=10):
    registri_read = [
        0x0445, 0x0446, 0x0447, 0x0448, 0x0449,
        0x044a, 0x044b, 0x044c, 0x0470, 0x0471
    ]
    expected_lengths = {14, 20}

    def connect_client():
        if proto == "RTU":
            return ModbusSerialClient(
                port=porta_com,
                baudrate=9600,
                # timeout=1,
                # stopbits=1,
                # bytesize=8,
                # parity='N'
            ), int(slave_id_rtu)
        elif proto == "TCP":
            return ModbusTcpClient(ip_tcp, port=8899, timeout=1), 1
        elif proto == "AzzurroHUB":
            return ModbusTcpClient(ip_hub, port=55400, timeout=5), int(slave_id_azzurro)
        else:
            raise ValueError(f"Protocollo non supportato: {proto}")

    for attempt_sn in range(max_retries):  # retry per intero SN
        sn_str = ""
        client, slave_id = connect_client()

        if not client.connect():
            print(f"[ERRORE] Connessione fallita ({proto})")
            return "UNKNOWN"

        try:
            for reg in registri_read:
                val_ok = False
                for _ in range(max_retries):  # retry per singolo registro
                    rr = client.read_holding_registers(address=reg, count=1, slave=slave_id)
                    if not rr.isError():
                        val = rr.registers[0]
                        hex_str = f"{val:04x}"
                        if hex_str == '0000':
                            val_ok = True
                            break
                        sn_str += chr(int(hex_str[0:2], 16))
                        sn_str += chr(int(hex_str[2:4], 16))
                        val_ok = True
                        break
                    # piccolo delay prima del retry
                    if proto == "AzzurroHUB":
                        time.sleep(3)  # attesa obbligatoria per non farsi bloccare dall'HUB
                    else:
                        time.sleep(0.5)  # attesa più breve per TCP/RTU

                if not val_ok:
                    print(f"[WARN] Registro {hex(reg)} non letto dopo {max_retries} tentativi")

        except Exception as e:
            print(f"[ERRORE lettura SN] {e}")
            sn_str = ""
        finally:
            client.close()

        sn_str = sn_str.strip()
        if len(sn_str) in expected_lengths:
            return sn_str  # SN valido trovato

        print(f"[WARN] SN incompleto (len={len(sn_str)}), retry {attempt_sn+1}/{max_retries}")

    return "UNKNOWN"


# --- Lettura DB inverter ---
def load_inverter_db(model, sn):
    db_path = os.path.join("./database", f"{model}.xlsx")
    if not os.path.exists(db_path):
        return None

    wb = openpyxl.load_workbook(db_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip() == str(sn).strip():
            return headers, row
    return None


# Funzione che crea una nuova finestra con messaggio HelloWorld
def open_subpanel(title):
    sub_win = tk.Toplevel(root)
    sub_win.title(title)
    label = tk.Label(sub_win, text="HelloWorld", font=("Arial", 16))
    label.pack(padx=20, pady=20)


# apro il thread per il log
def start_logging_routine(protocol, inverters, registers, file_path, sampling_time, total_time, shared_ins=None):
    global logging_running, logging_paused, rt_columns
    logging_running = True
    logging_paused = False

    def log_loop():
        global rt_columns
        start_time = time.time()
        # Prepara header CSV
        header = ["timestamp"]
        col_names = []
        for inv_index, inv in enumerate(inverters):
            for label, _, _ in registers:
                col_names.append(f"Inverter{inv_index+1}_{label}")
        header += col_names
        try:
            # assicura che la cartella esista
            from datetime import datetime
            outdir = os.path.dirname(file_path)
            if outdir:
                os.makedirs(outdir, exist_ok=True)
            # line-buffered, UTF-8, newline corretto per CSV
            with open(file_path, mode='w', newline='', encoding='utf-8', buffering=1) as f:
                writer = csv.writer(f)
                writer.writerow(header)
                f.flush()
                warned_ports = set()
                warned_file_lock = False
                while time.time() - start_time < total_time:
                    if not logging_running:
                        break
                    if logging_paused:
                        time.sleep(0.5); continue
                    timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    row_vals = []
                    # for inv in inverters:
                    #     address = inv["ip"]
                    #     modbus_id = int(inv["modbus"])
                    #     client = None
                    #     try:
                    #         if protocol == "RTU":
                    #             client = ModbusSerialClient(
                    #                 port=address, baudrate=9600, timeout=1, stopbits=1, bytesize=8, parity='N',
                    #             )
                    #         else:
                    #             port = 8899 if protocol == "TCP" else 55400
                    #             client = ModbusTcpClient(address, port=port, timeout=1)
                    #
                    #         if not client.connect():
                    #             raise ConnectionError(f"Connessione fallita su {address}")
                    #
                    #         reg_values = []
                    #         for _, reg, scale in registers:
                    #             reg = int(reg, 16) if isinstance(reg, str) and reg.lower().startswith("0x") else int(reg)
                    #             scale = int(scale)
                    #             # pymodbus 3.8 → unit=
                    #             result = client.read_holding_registers(address=reg, count=1, slave=modbus_id)
                    #             if not result or not hasattr(result, "registers"):
                    #                 reg_values.append(None)  # usa None invece di "ERR" per grafico
                    #             else:
                    #                 reg_values.append(result.registers[0] * scale)
                    #
                    #         client.close()
                    #         row_vals.extend(reg_values)
                    #
                    #
                    #     except Exception as e:
                    #         row_vals.extend([None] * len(registers))
                    #         if address not in warned_ports:
                    #             warned_ports.add(address)
                    #             try:
                    #                 messagebox.showwarning("Porta non disponibile",
                    #                                        f"{address}: {e}\n(continuo con gli altri dispositivi)")
                    #             except Exception:
                    #                 print(f"[WARN] {address}: {e}")
                    #     finally:
                    #         try:
                    #             if client:
                    #                 client.close()
                    #         except Exception:
                    #             pass
                    for inv in inverters:
                        # usa il service condiviso per leggere (ruolo da 'slave' flag)
                        reg_values = []
                        role = "slave" if inv.get("slave") else "master"
                        try:
                            for _, reg, scale in registers:
                                # reg può essere "0x...." o int
                                out = shared_ins.inv_broadcast_read(reg, count=1, role=role) if shared_ins else None
                                if out and isinstance(out, dict) and out:
                                    regs = next(iter(out.values()))  # prima entry del dict
                                    raw = regs[0] if regs else None
                                else:
                                    raw = None
                                # decode 16 bit  scaling  two's complement (se necessario)
                                val = decode_u16_auto(raw, scale=scale, signed_hint_thresh=0xF000)
                                reg_values.append(val)
                        except Exception as e:
                            # in caso d'errore su un inverter, logga vuoti ma continua con gli altri
                            reg_values = [None] * len(registers)
                            print(f"[WARN] Lettura {role} fallita: {e}")
                        row_vals.extend(reg_values)
                    # Normalizza riga (None/NaN -> stringa vuota) e scrivi
                    def _cell(v):
                        if v is None: return ""
                        if isinstance(v, float) and math.isnan(v): return ""
                        return v
                    row = [timestamp_str] + [_cell(v) for v in row_vals]
                    try:
                        writer.writerow(row)
                        f.flush()  # flush ad ogni campione (puoi togliere se vuoi flush periodico)
                    except PermissionError as e:
                        # file probabilmente aperto in Excel: avvisa una sola volta, poi continua il log
                        if not warned_file_lock:
                             warned_file_lock = True
                             try:
                                 messagebox.showwarning("File bloccato",f"Non riesco a scrivere su:\n{file_path}\n\nMotivo: {e}\n. Chiudi il file se è aperto (e.g. Excel). Continuerò a tentare.")
                             except Exception:
                                 print(f"[WARN] CSV lock: {e}")
                    except Exception as e:
                    # altre eccezioni: non bloccare il logger
                        print(f"[WARN] writerow fallita: {e}")
                    # === PUSH nei buffer realtime ===
                    with rt_lock:
                        # inizializza colonne se vuoto
                        if not rt_columns:
                            rt_columns.clear(); rt_columns.extend(col_names)
                        # allinea lunghezza
                        rt_time.append(timestamp_str)
                        for cname, val in zip(col_names, row_vals):
                            # numerico → float; altrimenti NaN (mantiene cardinalità uguale a rt_time)
                            try:
                                v = float(val)
                            except (TypeError, ValueError):
                                v = math.nan
                            rt_data[cname].append(v)
                    time.sleep(sampling_time)
            #messagebox.showinfo("Logging completato", f"File salvato:\n{file_path}")
            print(f"[INFO] Logging completato. File salvato: {file_path}")

            # Esporta anche in XLSX con fogli per inverter (nome = seriale)
            try:
                import pandas as pd, re
                from datetime import datetime
                xlsx_path = os.path.splitext(file_path)[0] + ".xlsx"
                df_all = pd.read_csv(file_path)
                with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as wr:
                    for idx, inv in enumerate(inverters, start=1):
                        prefix = f"Inverter{idx}_"
                        serial = inv.get("sn","INV"+str(idx))
                        # Excel sheet name constraints
                        safe_serial = re.sub(r'[:\\/?*\[\]]', "_", str(serial))[:31] or f"INV{idx}"
                        cols = [c for c in df_all.columns if c.startswith(prefix)]
                        if not cols:
                            continue
                        df_sheet = df_all[["timestamp"] + cols].copy()
                        # rinomina rimuovendo il prefisso
                        df_sheet.columns = ["timestamp"] + [c[len(prefix):] for c in cols]
                        df_sheet.to_excel(wr, sheet_name=safe_serial, index=False)
                print(f"[INFO] XLSX con fogli per inverter salvato: {xlsx_path}")
            except Exception as e:
                print(f"[WARN] esportazione XLSX per inverter fallita: {e}")
                # ====== Export sheet "LogErrori" per ciascun inverter ======
            try:
                import pandas as pd, re
                from datetime import datetime
                from openpyxl import load_workbook

                def _bcd4(n: int):
                    """Ritorna 4 nibble (0..9) da un U16: [d3,d2,d1,d0]."""
                    a = str(hex(n)).split('x')[1]
                    return [a[0], a[1], a[2], a[3]]

                def _safe_dt(y, m, d, hh, mm, ss):
                    try:
                        # anno su 2 cifre → 2000+YY (adatta se serve 20xx o 19xx)
                        Y = 2000 + y
                        return datetime(Y, m, d, hh, mm, ss)
                    except Exception:
                        return None

                # intervallo del test = durata del logging
                log_start_dt = datetime.fromtimestamp(start_time)
                log_end_dt = datetime.now()

                xlsx_path = os.path.splitext(file_path)[0] + ".xlsx"
                # se esiste già, apri in append; altrimenti crea nuovo file
                if os.path.isfile(xlsx_path):
                    writer_ctx = pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace")
                else:
                    writer_ctx = pd.ExcelWriter(xlsx_path, engine="openpyxl")

                with writer_ctx as wr:
                    for idx, inv in enumerate(inverters, start=1):
                        role = "slave" if inv.get("slave") else "master"
                        serial = str(inv.get("sn", f"INV{idx}"))
                        safe_serial = re.sub(r'[:\\/?*\[\]]', "_", serial)[:31] or f"INV{idx}"
                        sheet_name = f"{safe_serial}_LogErrori"

                        rows = []
                        # leggi fino a 100 eventi (0..99)
                        for k in range(10):
                            base = 0x1480 + 4 * k
                            try:
                                out = shared_ins.inv_broadcast_read(base, count=4,
                                                                    role=role) if shared_ins else None
                                # 'out' atteso: dict {<id> : [r0,r1,r2,r3]}
                                if not out or not isinstance(out, dict) or not out:
                                    continue
                                regs = next(iter(out.values()))
                                if not regs or len(regs) < 4:
                                    continue
                                code = regs[0] & 0xFFFF
                                y1, y2, m1, m2 = _bcd4(regs[1])
                                try:
                                    year = '0x' + y1 + y2
                                except:
                                    year = '0x' + y2
                                try:
                                    month = '0x' + m1 + m2
                                except:
                                    month = '0x' + m2
                                d1, d2, h1, h2 = _bcd4(regs[2])
                                try:
                                    day = '0x' + d1 + d2
                                except:
                                    day = '0x' + d2
                                try:
                                    hour = '0x' + h1 + h2
                                except:
                                    hour = '0x' + h2
                                M1, M2, S1, S2 = _bcd4(regs[3])
                                try:
                                    minute = '0x' + M1  + M2
                                except:
                                    minute = '0x' + M2
                                try:
                                    sec = '0x' + S1 + S2
                                except:
                                    sec = '0x' + S2
                                ts = _safe_dt(int(year, base=16), int(month, base=16), int(day, base=16),
                                              int(hour, base=16), int(minute, base=16), int(sec, base=16))
                                if not ts:
                                    continue
                                # filtra eventi del test
                                if log_start_dt <= ts <= log_end_dt:
                                    rows.append({
                                        "timestamp": ts.strftime("%Y-%m-%d %H:%M:%S"),
                                        "code_dec": int(code),
                                        "code_hex": f"0x{code:04X}",
                                        "fault_index": None,
                                        "register": None,
                                        "source": "HIST",
                                        "YY": year, "MM": month, "DD": day,
                                        "hh": hour, "mm": minute, "ss": sec
                                    })
                            except Exception as e:
                                print(f"[WARN] lettura LogErrori {role} evt#{k}: {e}")

                        # # ====== Errori attuali (ongoing) 0x0405..0x040E ======
                        # try:
                        #     # timestamp = ultima riga del CSV (o adesso)
                        #     try:
                        #         df_all_csv = pd.read_csv(file_path)
                        #         last_ts_str = str(
                        #             df_all_csv["timestamp"].iloc[-1]) if not df_all_csv.empty else None
                        #     except Exception:
                        #         last_ts_str = None
                        #     if not last_ts_str:
                        #         last_ts_str = log_end_dt.strftime("%Y-%m-%d %H:%M:%S")
                        ### TO DO: mappa dei bit per leggere gli errori.
                        #     cur = shared_ins.inv_broadcast_read(0x0405, count=1,
                        #                                         role=role) if shared_ins else None
                        #     if cur and isinstance(cur, dict) and cur:
                        #         regs = next(iter(cur.values()))
                        #         if regs and len(regs) >= 1:
                        #
                        #             for j, val in enumerate(regs[:10]):
                        #                 v = int(val)
                        #                 if v != 0:  # logga solo fault presenti
                        #                     rows.append({
                        #                         "timestamp": last_ts_str,
                        #                         "code_dec": v,
                        #                         "code_hex": f"0x{v:04X}",
                        #                         "fault_index": j + 1,
                        #                         "register": f"0x{0x0405 + j:04X}",
                        #                         "source": "ONGOING",
                        #                         "YY": "", "MM": "", "DD": "",
                        #                         "hh": "", "mm": "", "ss": ""
                        #                     })
                        # except Exception as e:
                        #     print(f"[WARN] lettura Fault1..10 {role}: {e}")

                        if rows:
                            df_err = pd.DataFrame(rows).sort_values("timestamp")
                            df_err.to_excel(wr, sheet_name=sheet_name, index=False)
                            print(f"[INFO] LogErrori scritto: {xlsx_path} [{sheet_name}] ({len(rows)} eventi)")
                        else:
                            print(f"[INFO] Nessun evento nel range per {safe_serial}")
            except Exception as e:
                print(f"[WARN] export LogErrori fallito: {e}")
                # ====== Report (se abbiamo un template singolo) ======
            try:
                from drivers.report_html import render_mppt_report_html
                tpl = current_report_ctx.get("template_path")
                serials = current_report_ctx.get("serials", [])
                if tpl and os.path.isfile(xlsx_path):
                    out_html = os.path.join(os.path.dirname(xlsx_path),
                                            f"Report_{os.path.splitext(os.path.basename(tpl))[0]}_{(serials[0] if serials else 'INV')}.html")
                    out_pdf = out_html.replace(".html", ".pdf")
                    print('inizio render')
                    _, pdf_path, graph_html = render_mppt_report_html(
                        log_xlsx_path=xlsx_path,
                        inverter_serials=serials,
                        template_path=tpl,
                        out_html_path=out_html,
                        out_pdf_path=out_pdf, # out_pdf_path=os.path.splitext(out_html)[0] + ".pdf"
                        logo_path="./misc/logo/logo.jpg",  # o None -> auto-pick
                        header_path=None, footer_path=None,
                        meta={"company": "GID Lab"}
                    )
                    print('fine render')
                    print("[INFO] Report HTML:", out_html, "| PDF:", pdf_path or "(non creato)")

            except Exception as e:
                print(f"[WARN] generazione report fallita: {e}")

            # (opzionale) piccolo toast non-modale in basso a destra
            # try:
            #     import tkinter as tk
            #     root = tk._default_root
            #     if root:
            #         toast = tk.Toplevel(root)
            #         toast.overrideredirect(True)
            #         toast.attributes("-topmost", True)
            #         lbl = tk.Label(toast, text=f"Log completato:\n{file_path}",
            #                        bg="#323232", fg="white", padx=12, pady=8, justify="left")
            #         lbl.pack()
            #         toast.update_idletasks()
            #         x = root.winfo_rootx() + root.winfo_width() - toast.winfo_width() - 20
            #         y = root.winfo_rooty() + root.winfo_height() - toast.winfo_height() - 20
            #         toast.geometry(f"+{x}+{y}")
            #         toast.after(2500, toast.destroy)  # chiudi dopo 2.5s
            # except Exception as _e:
            #     # se il toast fallisce, ignoriamo: il print sopra resta
            #     pass


        except Exception as e:
            messagebox.showerror("Errore logging", str(e))

    t = threading.Thread(target=log_loop, daemon=True)
    t.start()
    return t


# Gestisco il Log del test automatico
def open_log_panel():
    log_win = tk.Toplevel()
    log_win.title("Lancio Log Strumenti")
    log_win.minsize(850, 600)

    playlist_path_var = tk.StringVar(master=log_win, value="")  # percorso file .txt della playlist (vuoto = singolo test)

    protocol_var = tk.StringVar(value="RTU")
    ttk.Label(log_win, text="Protocollo di Comunicazione").pack(anchor="w", padx=10, pady=(10, 0))
    ttk.Combobox(log_win, textvariable=protocol_var, values=["TCP", "RTU", "AzzurroHUB"]).pack(anchor="w", padx=10)

    # Fino a 10 inverter
    inverter_entries = []

    inverter_frame = tk.LabelFrame(log_win, text="Inverter", font=("Arial", 10, "bold"))
    inverter_frame.pack(fill="x", pady=5)

    # Intestazioni colonna (2 blocchi da 5 inverter)
    for block in range(2):  # 2 blocchi di 5 inverter
        col_offset = block * 4
        headers = ["#", "Modbus ID", "IP / COM", "M/S"]
        for idx, header in enumerate(headers):
            tk.Label(
                inverter_frame,
                text=header,
                font=("Arial", 9, "bold")
            ).grid(row=0, column=col_offset + idx, padx=5, pady=2)

    # Righe inverter
    for i in range(10):
        block = i // 5
        row = (i % 5) + 1
        col_offset = block * 4

        # Numero inverter
        tk.Label(inverter_frame, text=f"{i + 1}").grid(row=row, column=col_offset, padx=5, pady=2)

        # Modbus ID
        modbus_entry = tk.Entry(inverter_frame, width=5)
        modbus_entry.grid(row=row, column=col_offset + 1, padx=5, pady=2)

        # IP / COM
        ip_entry = tk.Entry(inverter_frame, width=15)
        ip_entry.grid(row=row, column=col_offset + 2, padx=5, pady=2)

        # M/S checkbox
        slave_var = tk.BooleanVar()
        slave_cb = tk.Checkbutton(inverter_frame, variable=slave_var)
        slave_cb.grid(row=row, column=col_offset + 3, padx=5, pady=2)

        # # Alimentatore dropdown
        # alim_var = tk.StringVar(value="Nessuno")
        # alim_menu = ttk.Combobox(
        #     inverter_frame,
        #     textvariable=alim_var,
        #     values=["Nessuno", "DC1", "DC2", "DC3"],
        #     width=8,
        #     state="readonly"
        # )
        # alim_menu.grid(row=row, column=col_offset + 4, padx=5, pady=2)

        inverter_entries.append((modbus_entry, ip_entry, slave_var))#, alim_var))

    # === Registri di default ===
    default_registers = [
        ("System State", "0x0404", "1"),
        ("Active Output Power [kW]", "0x0485", "0.01"),
        ("Reactive Output Power [kVAr]", "0x0486", "0.01"),
        ("Apparent Output Power [kVA]", "0x0487", "0.01"),
        ("Active PCC Power [kW]", "0x0488", "0.01"),
        ("Reactive PCC Power [kVAr]", "0x0489", "0.01"),
        ("Apparent PCC Power [kVA]", "0x048A", "0.01"),
        ("Voltage DC1 [V]", "0x0584", "0.1"),
        ("Current DC1 [A]", "0x0585", "0.01"),
        ("Power DC1 [kW]", "0x0586", "0.01"),
        ("Voltage DC2 [V]", "0x0587", "0.1"),
        ("Current DC2 [A]", "0x0588", "0.01"),
        ("Power DC2 [kW]", "0x0589", "0.01"),
        ("Charge/Discharge Power [kW]", "0x0667", "0.1"),
        ("Battery SOC [%]", "0x0668", "1")
    ]

    # Sezione registri (20 registri in 4 colonne da 5 righe)
    registers_frame = tk.LabelFrame(log_win, text="Registri", font=("Arial", 10, "bold"))
    registers_frame.pack(fill="x", pady=5)
    # tk.Label(log_win, text="Registri da leggere", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10, 0))
    # registers_frame = tk.Frame(log_win)
    # registers_frame.pack(anchor="w", padx=10)

    register_entries = []

    # Intestazioni per ciascuna colonna
    for col in range(4):  # 4 colonne
        tk.Label(registers_frame, text="Label", font=("Arial", 9, "bold")).grid(row=0, column=col * 3, padx=2, pady=2)
        tk.Label(registers_frame, text="Registro", font=("Arial", 9, "bold")).grid(row=0, column=col * 3 + 1, padx=2,
                                                                                   pady=2)
        tk.Label(registers_frame, text="Scaling", font=("Arial", 9, "bold")).grid(row=0, column=col * 3 + 2, padx=2,
                                                                                  pady=2)

    # Registri (20 = 4 colonne × 5 righe)
    for i in range(20):
        col = i // 5
        row = (i % 5) + 1  # +1 per lasciare spazio alle intestazioni

        name_entry = tk.Entry(registers_frame, width=20)
        reg_entry = tk.Entry(registers_frame, width=10)
        scale_entry = tk.Entry(registers_frame, width=6)

        if i < len(default_registers):
            name, reg, scale = default_registers[i]
            name_entry.insert(0, name)
            reg_entry.insert(0, reg)
            scale_entry.insert(0, scale)
        else:
            scale_entry.insert(0, "1")

        name_entry.grid(row=row, column=col * 3, padx=2, pady=2)
        reg_entry.grid(row=row, column=col * 3 + 1, padx=2, pady=2)
        scale_entry.grid(row=row, column=col * 3 + 2, padx=2, pady=2)

        register_entries.append((name_entry, reg_entry, scale_entry))

    # File CSV
    # file_frame = tk.Frame(log_win)
    # file_frame.pack(anchor="w", padx=10, pady=(10, 0))
    # tk.Label(file_frame, text="Percorso file CSV:", font=("Arial", 10, "bold")).pack(side="left")
    # file_entry = tk.Entry(file_frame, width=60)
    # file_entry.pack(side="left", padx=5)
    # def browse_file():
    #     filepath = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
    #     if filepath:
    #         file_entry.delete(0, tk.END)
    #         file_entry.insert(0, filepath)
    # tk.Button(file_frame, text="Sfoglia", command=browse_file).pack(side="left")

    # Tempi
    time_frame = tk.Frame(log_win)
    time_frame.pack(anchor="w", padx=10, pady=10)
    # tk.Label(time_frame, text="Campionamento [s]:", font=("Arial", 10, "bold")).pack(side="left")
    # sampling_entry = tk.Entry(time_frame, width=5)
    # sampling_entry.insert(0, "5")
    # sampling_entry.pack(side="left", padx=5)
    # tk.Label(time_frame, text="Durata test [s]:", font=("Arial", 10, "bold")).pack(side="left", padx=(20, 0))
    # duration_entry = tk.Entry(time_frame, width=5)
    # duration_entry.insert(0, "60")
    # duration_entry.pack(side="left", padx=5)

    # Campionamento scelto dall'utente
    tk.Label(time_frame, text="Campionamento [s]:", font=("Arial", 10, "bold")).pack(side="left")
    global sampling_entry
    sampling_entry = tk.Entry(time_frame, width=5)
    sampling_entry.insert(0, "1")
    sampling_entry.pack(side="left", padx=5)

    # # Durata test calcolata automaticamente (readonly)
    # tk.Label(time_frame, text="Durata test [s]:", font=("Arial", 10, "bold")).pack(side="left", padx=(20, 0))
    # global duration_entry_var, duration_entry
    # duration_entry_var = tk.StringVar(value="60")
    # duration_entry = tk.Entry(time_frame, width=7, state="readonly", textvariable=duration_entry_var)
    # duration_entry.pack(side="left", padx=5)

    # Pulsanti
    def on_send_log():
        global logging_thread, logging_running, test_thread, current_shared_ins
        # ferma eventuale logger precedente e libera COM
        stop_logging_and_release()
        # file_path = file_entry.get().strip()
        # if not file_path:
        #     messagebox.showwarning("Attenzione", "Inserisci un percorso valido per il file CSV.")
        #     return

        # 1) Campionamento scelto dall'utente
        try:
            sampling = float(sampling_entry.get())
        except Exception:
            messagebox.showerror("Errore", "Campionamento non valido")
            return

        inverter_data = []
        for idx, (modbus_entry, ip_entry, slave_var) in enumerate(inverter_entries):
            ip_val = ip_entry.get().strip()
            modbus_val = modbus_entry.get().strip()
            if not ip_val or not modbus_val:
                continue
            mode = protocol_var.get().strip()
            if mode == "TCP":
                sn = read_SN(mode, ip_tcp=ip_val)
            elif mode == "RTU":
                sn = read_SN(mode, porta_com=ip_val, slave_id_rtu=int(modbus_val))
            elif mode == "AzzurroHUB":
                sn = read_SN(mode, ip_hub=ip_val, slave_id_azzurro=int(modbus_val))
            else:
                sn = "UNKNOWN"
            inverter_data.append({
                "sn": sn,
                "ip": ip_val,
                "modbus": int(modbus_val),
                "slave": slave_var.get(),
                #"alimentatore": alim_var.get()
            })

        print("[INFO] Inverter rilevati:", inverter_data)
        sn_for_test = next((x["sn"] for x in inverter_data if not x["slave"]), None)
        if not sn_for_test and inverter_data:
            sn_for_test = inverter_data[0]["sn"]

        from datetime import datetime
        session_dir = os.path.join("./Data", f"{sn_for_test}_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        os.makedirs(session_dir, exist_ok=True)

        # ignoriamo il file chooser: salviamo nella sessione con nome del test
        selected_template = template_var.get().strip()
        test_name = os.path.splitext(selected_template)[0] if selected_template and selected_template.lower() != "custom" else "custom"
        file_path = os.path.join(session_dir, f"{test_name}.csv")

        registers = []
        for name_entry, reg_entry, scale_entry in register_entries:
            name = name_entry.get().strip()
            reg = reg_entry.get().strip()
            scale = scale_entry.get().strip()
            if name and reg and scale:
                registers.append((name, reg, scale))

        if not registers:
            messagebox.showerror("Errore", "Nessun registro valido selezionato.")
            return

        # # --- Durata logging = durata test (stima automatica) -----------------
        # # Regole: custom/altro = somma 'tempo'; curva MPPT = tempo[0] * (Vmax/5)
        def _estimate_template_duration(tpl_path: str, sn: str) -> float:
            try:
                import pandas as pd
                df_t = pd.read_excel(tpl_path)
            except Exception as e:
                print(f"[WARN] Impossibile leggere template {tpl_path}: {e}")
                return 0.0

            base = os.path.basename(tpl_path)
            # MPPT dal nome file (coerente con il resto del codice)
            if "curva MPPT" in base:
                # step (s): prima cella di 'tempo', default 1.0
                try:
                    step_s = float(str(df_t["tempo"].iloc[0]).replace(",", "."))
                except Exception:
                    step_s = 1.0
                # Vmax dal DB modello (via SN)
                vmax = 0.0
                try:
                    info = parse_sn(sn) if sn else {}
                    family = (info.get("family",[None])[0] if isinstance(info.get("family"), (list,tuple)) else info.get("family"))
                    model_code = (info.get("model_code",[None])[0] if isinstance(info.get("model_code"), (list,tuple)) else info.get("model_code"))
                    if family and model_code:
                        db_path = os.path.join("./database", f"{family}.xlsx")
                        df_inv = pd.read_excel(db_path, dtype={"Unnamed: 0": str}, keep_default_na=False)
                        df_inv["Unnamed: 0"] = (df_inv["Unnamed: 0"].astype(str).str.strip()
                                                .str.replace(r"\.0+$", "", regex=True).str.zfill(3))
                        key = str(model_code).zfill(3)
                        vmax = float(df_inv.loc[df_inv["Unnamed: 0"] == key, "MAX V"].iloc[0])
                except Exception as e:
                    print(f"[WARN] stima Vmax fallita: {e}")
                    vmax = 100.0  # fallback prudente
                return float(step_s) * (float(vmax) / 5.0) + 60 #+60 sono i secondi per la preconnessione nel caso in cui fosse spento

            elif "MAX SOUT" in base:
                try:
                    step_s = float(str(df_t["tempo"].iloc[0]).replace(",", "."))
                except Exception:
                    step_s = 1.0
                return 1800 * step_s / 45

            elif "0-INJ" in base:
                try:
                    step_s = float(str(df_t["tempo"].iloc[0]).replace(",", "."))
                except Exception:
                    step_s = 1.0
                return 15 * step_s

            # default/custom/altro: somma dei tempi
            try:
                import pandas as pd
                if "tempo" in df_t.columns:
                    tempi = pd.to_numeric(df_t["tempo"].astype(str).str.replace(",", "."), errors="coerce").fillna(0.0)
                    return float(tempi.sum())
            except Exception as e:
                print(f"[WARN] somma tempi fallita su {tpl_path}: {e}")
            return 0.0
        # ---------------------------------------------------------------------

        # Calcolo durata in base a playlist o singolo template
        try:
            playlist = playlist_path_var.get().strip()
        except Exception:
            playlist = ""

        if playlist and os.path.isfile(playlist):
            # Somma durate di tutti i template della playlist
            total_dur = 0.0
            try:
                with open(playlist, "r", encoding="utf-8") as f:
                    names = [ln.strip() for ln in f if ln.strip() and not ln.strip().startswith("#")]
                for name in names:
                    cand = os.path.join(template_folder, f"{name}.xlsx")
                    if os.path.isfile(cand):
                        total_dur += _estimate_template_duration(cand, sn_for_test)
                # piccolo margine di coda
                if total_dur > 0:
                    duration = total_dur + 5.0
            except Exception as e:
                print(f"[WARN] stima durata playlist fallita: {e}")
        else:
            # Singolo template selezionato
            selected_template = template_var.get().strip().split('.xlsx')[0]
            candidate = os.path.join(template_folder, f"{selected_template}.xlsx")
            if os.path.isfile(candidate):
                est = _estimate_template_duration(candidate, sn_for_test)
                if est > 0:
                    duration = est
        # --- fine stima durata ------------------------------------------------

        # 0) Contesto report (singolo test: salviamo il template; playlist: lo lasciamo vuoto)
        try:
            selected_template = template_var.get().strip()
            candidate = os.path.join(template_folder, f"{selected_template}.xlsx")
            template_path = candidate if os.path.isfile(candidate) else None
        except Exception:
            template_path = None
        serials_for_report = [d.get("sn","INV") for d in inverter_data]
        # Se playlist, lo lasciamo vuoto (gestiremo per-test più avanti)
        if playlist and os.path.isfile(playlist):
            current_report_ctx.clear()
        else:
            current_report_ctx.clear()
            if template_path:
                current_report_ctx.update({"template_path": template_path, "serials": serials_for_report})

        # 1) Crea service condiviso ma SOLO per le porte realmente presenti
        inv_cfgs = build_inv_cfgs_from_ui(protocol_var.get(), inverter_data)

        dc_fixed = {"DC1": "ASRL20::INSTR", "DC2": "ASRL21::INSTR", "DC3": "ASRL22::INSTR"}
        present_dc = {}
        for ch, addr in dc_fixed.items():
            ok, info = _visa_is_present(addr)
            if ok:
                present_dc[ch] = addr
                print(f"[DC] {ch} OK @ {addr}" + (f" — {info}" if info else ""))
            else:
                print(f"[DC] {ch} NON presente @ {addr}")

        ac_fixed = "ASRL5::INSTR"
        ac_ok, ac_info = _visa_is_present(ac_fixed)
        ac_addr = ac_fixed if ac_ok else None
        print(f"[AC] {'OK' if ac_ok else 'NON presente'} @ {ac_fixed}" + (f" — {ac_info}" if ac_ok and ac_info else ""))

        current_shared_ins = Instruments(
            dc_map=(present_dc if present_dc else None),
            ac_addr=ac_addr,
            inv_cfgs=inv_cfgs,
            protocol=protocol_var.get()
        )
        # 2) Avvia logging con service condiviso e conserva il thread
        logging_thread = start_logging_routine(protocol_var.get(), inverter_data, registers, file_path, sampling,
                                               duration, shared_ins=current_shared_ins)
        # dopo aver popolato inverter_data e registers e avviato logging_thread
        # ricostruisci i nomi colonna come nel logger:
        col_names = []
        for inv_index, inv in enumerate(inverter_data):
            for label, _, _ in registers:
                col_names.append(f"Inverter{inv_index + 1}_{label}")

        # default: prima grandezza = Inverter1_<primo label>
        default_col = col_names[0] if col_names else None
        open_realtime_panel(col_names, default_col=default_col)

        # 3) Singolo test o playlist
        playlist = playlist_path_var.get().strip()
        if playlist and os.path.isfile(playlist):
            # ESECUZIONE IN SERIE: per non bloccare la UI, lancia in un thread dedicato
            # def _run_playlist():
            #     try:
            #         run_tests_playlist(
            #             playlist_path=playlist,
            #             sn=sn_for_test,
            #             protocol=protocol_var.get(),
            #             inverter_data=inverter_data,
            #             shared_ins=current_shared_ins,
            #             template_folder=template_folder,
            #         )
            #         print("[INFO] Playlist completata.")
            #     except Exception as e:
            #         try:
            #             messagebox.showerror("Errore playlist", str(e))
            #         except Exception:
            #             print(f"[ERR] Playlist: {e}")

            def run_playlist_segmented(playlist_path, sn, protocol, inverter_data, shared_ins, template_folder,
                                       sampling, registers, session_dir):
                # legge i nomi dei test
                with open(playlist_path, "r", encoding="utf-8") as f:
                    names = [ln.strip() for ln in f if ln.strip() and not ln.strip().startswith("#")]

                for name in names:
                    template_path = os.path.join(template_folder, f"{name}.xlsx")
                    if not os.path.isfile(template_path):
                        print(f"[WARN] Template non trovato: {template_path}")
                        continue

                    # stima durata SOLO per questo test (riusa la tua _estimate_template_duration)
                    dur = _estimate_template_duration(template_path, sn) or 60.0

                    # file CSV dedicato in session_dir
                    csv_path = os.path.join(session_dir, f"{name}.csv")

                    # avvia logger per questo test
                    log_thread = start_logging_routine(protocol, inverter_data, registers, csv_path, sampling, dur,
                                                       shared_ins=shared_ins)

                    # avvia test singolo
                    t = run_test_from_template(template_path, sn, protocol, inverter_data, shared_ins=shared_ins)

                    # attendi fine test + log
                    if t: t.join()
                    if log_thread: log_thread.join()

                    # qui il logger ha già convertito in XLSX e (se configurato) creato il report
                    print(f"[INFO] Test '{name}' completato. Dati in: {csv_path}")

            def _run_playlist_segmentato():
                try:
                    # ricostruisci i registers dalla UI una volta sola (già li hai in 'registers')
                    run_playlist_segmented(
                        playlist_path=playlist,
                        sn=sn_for_test,
                        protocol=protocol_var.get(),
                        inverter_data=inverter_data,
                        shared_ins=current_shared_ins,
                        template_folder=template_folder,
                        sampling=float(sampling_entry.get()),
                        registers=registers,
                        session_dir=session_dir
                    )
                    print("[INFO] Playlist completata:", session_dir)
                    # genera/aggiorna l'indice di sessione (HTML + PDF)
                    try:
                        from drivers.report_html import render_session_index
                        idx_html = os.path.join(session_dir, "index.html")
                        idx_pdf = os.path.join(session_dir, "index.pdf")
                        render_session_index(session_dir, out_html_path=idx_html, out_pdf_path=idx_pdf)
                        print("[INFO] Index scritto:", idx_html)
                    except Exception as e:
                        print("[WARN] index sessione non creato:", e)

                except Exception as e:
                    try:
                        messagebox.showerror("Errore playlist", str(e))
                    except Exception:
                        print(f"[ERR] Playlist: {e}")

            threading.Thread(target=_run_playlist_segmentato, daemon=True).start()
        else:
            # PLAYLIST se selezionata, altrimenti singolo test
            def _run_tests():
                playlist = playlist_path_var.get().strip() if 'playlist_path_var' in globals() else ""
                try:
                    if playlist and os.path.isfile(playlist):
                        run_tests_playlist(playlist, sn_for_test, protocol_var.get(), inverter_data,
                                           shared_ins=current_shared_ins, template_folder=template_folder)
                    else:
                        selected_template = template_var.get().strip().split('.xlsx')[0]
                        candidate = os.path.join(template_folder, f"{selected_template}.xlsx")
                        template_path = candidate if os.path.isfile(candidate) else "custom.xlsx"
                        run_test_from_template(template_path, sn_for_test, protocol_var.get(), inverter_data,
                                               shared_ins=current_shared_ins)
                except Exception as e:
                    try:
                        messagebox.showerror("Errore test", str(e))
                    except:
                        print("[ERR]", e)

            threading.Thread(target=_run_tests, daemon=True).start()

    def pause_logging():
        global logging_paused
        logging_paused = True

    def resume_logging():
        global logging_paused
        logging_paused = False

    button_frame = tk.Frame(log_win)
    button_frame.pack(fill="x", pady=5, padx=5)

    # --- Template Test ---
    tk.Label(button_frame, text="Seleziona template test:", font=("Arial", 10, "bold")).pack(side="left", padx=(0,5))

    # Caricamento file disponibili in ./template/
    template_folder = "./template/"
    template_files = []
    if os.path.exists(template_folder):
        template_files = [f for f in os.listdir(template_folder) if f.lower().endswith((".xlsx", ".xls"))]
    template_options = template_files

    template_var = tk.StringVar(value="custom.xlsx")
    template_menu = ttk.OptionMenu(button_frame, template_var, "custom.xlsx", *template_options)

    template_menu.pack(side="left", padx=(0, 20))

    #frame dei bottoni
    tk.Button(button_frame, text="Send", bg="lightgreen", command=on_send_log).pack(side="right", padx=2)
    tk.Button(button_frame, text="Stop Log", command=stop_logging_and_release).pack(side="right", padx=2)
    tk.Button(button_frame, text="Pause", bg="orange", command=pause_logging).pack(side="right", padx=2)
    tk.Button(button_frame, text="Resume", bg="lightblue", command=resume_logging).pack(side="right", padx=2)
    tk.Button(button_frame, text="Exit", bg="red", fg="white", command=log_win.destroy).pack(side="right", padx=2)

    # -- Riga playlist (.txt) --
    playlist_row = tk.Frame(log_win)
    # era: playlist_row.pack(side="left", padx=(0, 20))
    # meglio "top" + fill X per allinearlo bene in alto a sinistra
    playlist_row.pack(side="top", fill="x", padx=10, pady=2, anchor="w")

    tk.Label(playlist_row, text="Playlist (.txt):").pack(side="left", padx=(0, 6))
    playlist_entry = tk.Entry(playlist_row, textvariable=playlist_path_var, width=48, state="readonly")
    playlist_entry.pack(side="left", padx=(0, 6))

    # (ricorda di avere queste funzioni definite)
    # def _browse_playlist(): ...
    # def _clear_playlist(): ...

    def _browse_playlist():
        from tkinter import filedialog
        p = filedialog.askopenfilename(
            title="Seleziona playlist di test",
            filetypes=[("Playlist di test", "*.txt"), ("Tutti i file", "*.*")]
        )
        if p:
            playlist_path_var.set(p)

    def _clear_playlist():
        playlist_path_var.set("")

    tk.Button(playlist_row, text="Sfoglia…", command=_browse_playlist).pack(side="left", padx=(0, 6))
    tk.Button(playlist_row, text="X", width=2, command=_clear_playlist).pack(side="left")


    # Frame dinamico per configurazione test
    test_config_frame = tk.Frame(log_win)
    test_config_frame.pack(fill="x", padx=10, pady=10)

    def hide_test_config():
        for widget in test_config_frame.winfo_children():
            widget.destroy()

    def show_test_config():
        hide_test_config()

        # Titolo Configurazione Test
        tk.Label(test_config_frame, text="Configurazione Test", font=("Arial", 10, "bold")).pack(anchor="w", padx=10,
                                                                                                 pady=(10, 5))

        # Frame per i checkbox in linea
        config_frame = tk.Frame(test_config_frame)
        config_frame.pack(fill="x", padx=10, pady=5)

        dc_power_vars = []
        for idx in range(3):
            var = tk.BooleanVar()
            cb = tk.Checkbutton(config_frame, text=f"Alimentatore DC{idx + 1}", variable=var)
            cb.pack(side="left", padx=5)
            dc_power_vars.append(var)

        ac_var = tk.BooleanVar()
        tk.Checkbutton(config_frame, text="Usa simulatore AC", variable=ac_var).pack(side="left", padx=5)

        inv_var = tk.BooleanVar()
        tk.Checkbutton(config_frame, text="Controllo inverter sotto test", variable=inv_var).pack(side="left", padx=5)

    def on_template_change(*args):
        selected = template_var.get()
        if selected == "custom.xlsx":
            hide_test_config()
        else:
            # show_test_config()
            meretrice = 0

    template_var.trace_add("write", on_template_change)


# Gestisco l'inverter
def open_inverter_control_panel():
    inv_win = tk.Toplevel()
    inv_win.title("Controllo Inverter")
    inv_win.minsize(100,100)

    # Variabili
    protocollo = tk.StringVar(value="TCP")
    operazione = tk.StringVar(value="Lettura")
    registro = tk.StringVar()
    num_reg = tk.StringVar()
    scaling = tk.StringVar()
    value = tk.StringVar()
    porta_com = tk.StringVar()
    slave_id_rtu = tk.StringVar()
    ip_tcp = tk.StringVar()
    ip_hub = tk.StringVar()
    slave_id_azzurro = tk.StringVar()

    # Funzione invio comandi (mock)
    def send_command():
        try:
            proto = protocollo.get()
            op = operazione.get()
            reg = int(registro.get(), 16)
            num = int(num_reg.get())
            scale = int(scaling.get())
            result = None

            # === CLIENT CONFIG ===
            if proto == "RTU":
                client = ModbusSerialClient(
                    port=porta_com.get(),
                    baudrate=9600,
                    timeout=1,
                    stopbits=1,
                    bytesize=8,
                    parity='N'
                )
                slave_id = int(slave_id_rtu.get())
            elif proto == "TCP":
                client = ModbusTcpClient(ip_tcp.get(), port=8899, timeout=1)
                slave_id = 1  # Default slave
            elif proto == "AzzurroHUB":
                client = ModbusTcpClient(ip_hub.get(), port=55400, timeout=1)
                slave_id = int(slave_id_azzurro.get())
            else:
                raise ValueError("Protocollo non valido")

            # === CONNESSIONE ===
            if not client.connect():
                raise ConnectionError("Impossibile connettersi al dispositivo")

            if op == "Lettura":
                result = client.read_holding_registers(address=reg, count=num, slave=slave_id)
                if isinstance(result, ModbusIOException) or not result:
                    raise IOError("Errore nella lettura")
                values = [val * scale for val in result.registers]
                messagebox.showinfo("Risultato Lettura", f"Valori: {values}")

            elif op == "Scrittura":
                val_input = value.get()
                try:
                    val_list = ast.literal_eval(val_input)
                    if not isinstance(val_list, list):
                        val_list = [int(val_list)]
                except:
                    raise ValueError("Formato value non valido (usa [1,2,3] o singolo numero)")

                scaled_vals = [int(v / scale) for v in val_list]
                if len(scaled_vals) == 1:
                    client.write_register(reg, scaled_vals[0], slave=slave_id)
                else:
                    client.write_registers(reg, scaled_vals, slave=slave_id)
                messagebox.showinfo("Scrittura", f"Scritti: {val_list} (scalati: {scaled_vals})")

            client.close()

        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def exit_panel():
        inv_win.destroy()

    # Campi dinamici
    def update_fields(*args):
        frame_rtu.forget()
        frame_tcp.forget()
        frame_azzurro.forget()
        frame_value.forget()

        if protocollo.get() == "RTU":
            frame_rtu.pack(pady=5, fill="x")
        elif protocollo.get() == "TCP":
            frame_tcp.pack(pady=5, fill="x")
        elif protocollo.get() == "AzzurroHUB":
            frame_azzurro.pack(pady=5, fill="x")

        if operazione.get() == "Scrittura":
            frame_value.pack(pady=5, fill="x")

    # UI layout base
    ttk.Label(inv_win, text="Protocollo Comunicazione:").pack(pady=(10, 0))
    ttk.OptionMenu(inv_win, protocollo, "RTU", "TCP", "RTU", "AzzurroHUB").pack()

    ttk.Label(inv_win, text="Lettura o Scrittura?").pack(pady=(10, 0))
    ttk.OptionMenu(inv_win, operazione, "Lettura", "Lettura", "Scrittura").pack()

    ttk.Label(inv_win, text="Scrivi il registro in hex (es. 0x0010):").pack(pady=(10, 0))
    tk.Entry(inv_win, textvariable=registro).pack()

    ttk.Label(inv_win, text="# registri da leggere:").pack(pady=(10, 0))
    tk.Entry(inv_win, textvariable=num_reg).pack()

    ttk.Label(inv_win, text="Scaling:").pack(pady=(10, 0))
    tk.Entry(inv_win, textvariable=scaling).pack()

    # Frame value (Scrittura)
    frame_value = tk.Frame(inv_win)
    ttk.Label(frame_value, text="Value (array o stringa):").pack(pady=(0, 5))
    tk.Entry(frame_value, textvariable=value).pack()

    # Frame RTU
    frame_rtu = tk.Frame(inv_win)
    ttk.Label(frame_rtu, text="Porta COM (es. COM3):").pack(pady=(0, 5))
    tk.Entry(frame_rtu, textvariable=porta_com).pack()
    ttk.Label(frame_rtu, text="Indirizzo Modbus Inverter:").pack(pady=(10, 0))
    tk.Entry(frame_rtu, textvariable=slave_id_rtu).pack()

    # Frame TCP
    frame_tcp = tk.Frame(inv_win)
    ttk.Label(frame_tcp, text="IP dell'Inverter:").pack(pady=(0, 5))
    tk.Entry(frame_tcp, textvariable=ip_tcp).pack()

    # Frame Azzurro HUB
    frame_azzurro = tk.Frame(inv_win)
    ttk.Label(frame_azzurro, text="IP dell'HUB:").pack(pady=(0, 5))
    tk.Entry(frame_azzurro, textvariable=ip_hub).pack()
    ttk.Label(frame_azzurro, text="Indirizzo Modbus Inverter:").pack(pady=(10, 0))
    tk.Entry(frame_azzurro, textvariable=slave_id_azzurro).pack()

    # Pulsanti
    button_frame = tk.Frame(inv_win)
    button_frame.pack(pady=20)
    tk.Button(button_frame, text="Send", bg="lightgreen", command=send_command).grid(row=0, column=0, padx=10)
    tk.Button(button_frame, text="Exit", bg="red", fg="white", command=exit_panel).grid(row=0, column=1, padx=10)

    # Tracciamento dinamico
    protocollo.trace_add("write", update_fields)
    operazione.trace_add("write", update_fields)

    update_fields()  # inizializzazione


# Gestisco l'AC Source
def open_ac_control_panel():
    ac_win = tk.Toplevel()
    ac_win.title("Controllo AC")
    ac_win.minsize(100, 100)

    # Variabili
    selected_tipo = tk.StringVar(value="Monofase")
    vac = tk.StringVar()
    freq = tk.StringVar()

    inst_ac = None  # Oggetto pyvisa

    def send_command():
        nonlocal inst_ac
        try:
            vac_val = float(vac.get())
            freq_val = float(freq.get())
            tipo = selected_tipo.get()

            rm = pyvisa.ResourceManager()
            inst_ac = rm.open_resource("ASRL5::INSTR")  # <--- Adatta qui la porta!

            inst_ac.write(f'VOLT {vac_val}')
            inst_ac.write(f'FREQ {freq_val}')

            if tipo == "Monofase":
                inst_ac.write("SYST:FUNC ONE")
            else:
                inst_ac.write("SYST:FUNC THREE")

            messagebox.showinfo("Successo", "Comandi AC inviati correttamente.")

        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def turn_on():
        try:
            if inst_ac:
                inst_ac.write('OUTP ON')
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def turn_off():
        try:
            if inst_ac:
                inst_ac.write('OUTP OFF')
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def exit_panel():
        ac_win.destroy()

    # UI
    ttk.Label(ac_win, text="Seleziona la tipologia di rete:").pack(pady=5)
    ttk.OptionMenu(ac_win, selected_tipo, "Monofase", "Monofase", "Trifase").pack()

    ttk.Label(ac_win, text="Vrms [V]:").pack(pady=(10, 0))
    tk.Entry(ac_win, textvariable=vac).pack()

    ttk.Label(ac_win, text="Freq. [Hz]:").pack(pady=(10, 0))
    tk.Entry(ac_win, textvariable=freq).pack()

    # Pulsanti ON/OFF
    button_frame1 = tk.Frame(ac_win)
    button_frame1.pack(pady=(20, 5))
    tk.Button(button_frame1, text="Turn On", bg="lightgreen", command=turn_on).grid(row=0, column=0, padx=10)
    tk.Button(button_frame1, text="Turn Off", bg="red", fg="white", command=turn_off).grid(row=0, column=1, padx=10)

    # Pulsanti SEND/EXIT
    button_frame2 = tk.Frame(ac_win)
    button_frame2.pack(pady=5)
    tk.Button(button_frame2, text="Send", bg="lightgreen", command=send_command).grid(row=0, column=0, padx=10)
    tk.Button(button_frame2, text="Exit", bg="red", fg="white", command=exit_panel).grid(row=0, column=1, padx=10)


# Gestisco le DC Sources
def open_dc_control_panel():
    dc_win = tk.Toplevel()
    dc_win.title("Controllo DC")
    dc_win.minsize(100, 100)

    # Variabili
    porta_map = {
        "Strumento 1": "ASRL20::INSTR",
        "Strumento 2": "ASRL21::INSTR",
        "Strumento 3": "ASRL22::INSTR"
        # Aggiungi qui se aggiungi altri strumenti DC
    }
    selected_strumento = tk.StringVar(value="Strumento 1")
    voc = tk.StringVar()
    isc = tk.StringVar()
    ff = tk.StringVar(value="1.0")

    inst = None  # Variabile di riferimento per pyvisa

    def send_command():
        nonlocal inst
        try:
            voc_val = float(voc.get())
            isc_val = float(isc.get())
            ff_val = float(ff.get())

            if not (0 <= ff_val <= 1):
                raise ValueError("Fattore di forma deve essere tra 0 e 1")

            porta = porta_map[selected_strumento.get()]
            rm = pyvisa.ResourceManager()
            inst = rm.open_resource(porta)

            inst.write(f'SOL:USER:VOC {voc_val}')
            inst.write(f'SOL:USER:VMP {voc_val * ff_val}')
            inst.write(f'SOL:USER:ISC {isc_val}')
            inst.write(f'SOL:USER:IMP {isc_val * ff_val}')
            messagebox.showinfo("Successo", "Comandi inviati correttamente.")
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def turn_on():
        try:
            if inst:
                inst.write('OUTP 1')
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def turn_off():
        try:
            if inst:
                inst.write('OUTP 0')
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def exit_panel():
        dc_win.destroy()

    # UI Layout
    ttk.Label(dc_win, text="Seleziona lo strumento da controllare:").pack(pady=5)
    ttk.OptionMenu(dc_win, selected_strumento, *porta_map.keys()).pack()

    ttk.Label(dc_win, text="Tensione Voc [V]:").pack(pady=(10, 0))
    tk.Entry(dc_win, textvariable=voc).pack()

    ttk.Label(dc_win, text="Corrente Isc [A]:").pack(pady=(10, 0))
    tk.Entry(dc_win, textvariable=isc).pack()

    ttk.Label(dc_win, text="Fattore di forma (0 - 1):").pack(pady=(10, 0))
    tk.Entry(dc_win, textvariable=ff).pack()

    # Pulsanti Turn On / Turn Off
    button_frame1 = tk.Frame(dc_win)
    button_frame1.pack(pady=(20, 5))
    tk.Button(button_frame1, text="Turn On", bg="lightgreen", command=turn_on).grid(row=0, column=0, padx=10)
    tk.Button(button_frame1, text="Turn Off", bg="red", fg="white", command=turn_off).grid(row=0, column=1, padx=10)

    # Pulsanti Send / Exit (sullo stesso livello)
    button_frame2 = tk.Frame(dc_win)
    button_frame2.pack(pady=5)
    tk.Button(button_frame2, text="Send", bg="lightgreen", command=send_command).grid(row=0, column=0, padx=10)
    tk.Button(button_frame2, text="Exit", bg="red", fg="white", command=exit_panel).grid(row=0, column=1, padx=10)


# Callback per ogni pulsante
def on_dc_control():
    open_dc_control_panel()


def on_ac_control():
    open_ac_control_panel()


def on_inverter_control():
    open_inverter_control_panel()


def on_log_start():
    open_log_panel()


def on_test_start():
    open_subpanel("Lancio Test Automatici")


def on_realtime_monitor():
    open_subpanel("Monitoraggio Realtime")


def on_generate_report():
    import os, sys, subprocess, threading
    from tkinter import filedialog, messagebox, ttk
    import tkinter as tk
    import pandas as pd

    # 1) XLSX di log
    xlsx_path = filedialog.askopenfilename(
        title="Seleziona file XLSX di log",
        initialdir=os.path.join(".", "Data"),
        filetypes=[("Excel (*.xlsx)", "*.xlsx"), ("Tutti i file", "*.*")]
    )
    if not xlsx_path:
        return

    # 2) Template usato
    initdir = "./template"
    template_path = filedialog.askopenfilename(
        title="Seleziona il template usato",
        initialdir=os.path.join(".", "template"),
        filetypes=[("Template Excel (*.xlsx)", "*.xlsx"), ("Tutti i file", "*.*")]
    )
    if not template_path:
        return

    # # 3) Logo (opzionale)
    # logo_path = filedialog.askopenfilename(
    #     title="Seleziona un logo (opzionale)",
    #     filetypes=[("Immagini (PNG/JPG/PDF)", "*.png;*.jpg;*.jpeg;*.pdf"), ("Tutti i file", "*.*")]
    # )
    # if not logo_path:
    logo_path = None
    # # 4) Header/Footer (opzionali)
    # header_path = filedialog.askopenfilename(
    #     title="Seleziona header (PNG) opzionale",
    #     initialdir=os.path.join(".", "misc", "carta intestata"),
    #     filetypes=[("PNG", "*.png"), ("Tutti i file", "*.*")]
    # )
    # footer_path = filedialog.askopenfilename(
    #     title="Seleziona footer (PNG) opzionale",
    #     initialdir=os.path.join(".", "misc", "carta intestata"),
    #     filetypes=[("PNG", "*.png"), ("Tutti i file", "*.*")]
    # )
    header_path = None #header_path or None
    footer_path = None #footer_path or None

    # Seriali: deduciamo i fogli "dati" (quelli che NON finiscono con _LogErrori)
    try:
        xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
        serials = [s for s in xls.sheet_names if not s.endswith("_LogErrori")]
        if not serials:
            serials = [xls.sheet_names[0]]
    except Exception as e:
        messagebox.showerror("Errore", f"Impossibile leggere l'XLSX:\n{e}")
        return

    # PDF di output
    base_tpl = os.path.splitext(os.path.basename(template_path))[0]
    first_sn = serials[0] if serials else "INV"
    out_pdf = os.path.join(os.path.dirname(xlsx_path), f"Report_{base_tpl}_{first_sn}.pdf")

    # Genera (thread separato) con finestrella di progresso
    def _worker():
        try:
            from drivers.report_html import render_mppt_report_html
            out_html = os.path.join(os.path.dirname(xlsx_path),
                            f"Report_{os.path.splitext(os.path.basename(template_path))[0]}_{serials[0]}.html")
            out_html, pdf_path, _ = render_mppt_report_html(
                log_xlsx_path = xlsx_path,
                inverter_serials = serials,
                template_path = template_path,
                out_html_path = out_html,
                out_pdf_path = os.path.splitext(out_html)[0] + ".pdf",   # << aggiungi
                logo_path = logo_path,
                header_path = header_path,
                footer_path = footer_path,
                meta = {"company": "GID Lab"}
            )
            msg = f"Report HTML creato:\n{out_html}"
            if pdf_path: msg += f"\nPDF: {pdf_path}"
            # torna nel main-thread per notifiche/aperture
            root.after(0, lambda: (
                messagebox.showinfo("Report", msg),
                # apri SEMPRE l'HTML
                (os.startfile(out_html) if os.name == "nt" else subprocess.Popen(["open" if sys.platform.startswith("darwin") else "xdg-open", out_html])),
                # apri anche il PDF se esiste
                (os.startfile(pdf_path) if (pdf_path and os.name == "nt") else None)))
        except Exception as e:
            err_msg = f"Generazione report fallita:\n{e}"
            print(err_msg)
            root.after(0, lambda m=err_msg: messagebox.showerror("Errore", m))
    # progress non modale
    try:
        prog = tk.Toplevel(root)
        prog.title("Generazione report")
        tk.Label(prog, text="Sto generando il report...").pack(padx=12, pady=8)
        pb = ttk.Progressbar(prog, mode="indeterminate", length=260)
        pb.pack(padx=12, pady=(0, 10))
        pb.start(10)
        def _close_prog():
            try: pb.stop(); prog.destroy()
            except: pass
          # chiudi quando il worker termina (tramite callback nel main-thread)


        def _run_and_close():
            try: _worker()
            finally: root.after(0, _close_prog)
        threading.Thread(target=_run_and_close, daemon=True).start()
    except Exception:
        # se il progress fallisce, almeno lancia il worker
        threading.Thread(target=_worker, daemon=True).start()


def on_generate_session_report():
    import os, sys, subprocess
    from tkinter import filedialog, messagebox
    from drivers.report_html import render_mppt_report_html, render_session_index

    # scegli la cartella di sessione (default .\Data)
    sess_dir = filedialog.askdirectory(
        title="Seleziona cartella sessione (.\\Data\\SN_timestamp)",
        initialdir=os.path.join(".", "Data")
    )
    if not sess_dir:
        return
    if not os.path.isdir(sess_dir):
        messagebox.showerror("Errore", f"Cartella non valida:\n{sess_dir}")
        return

    # per ogni XLSX nella sessione prova a ricostruire il template e generare il report
    template_root = os.path.join(".", "template")
    xlsx_files = [f for f in os.listdir(sess_dir) if f.lower().endswith(".xlsx")]
    if not xlsx_files:
        messagebox.showwarning("Attenzione", "Nessun XLSX trovato nella cartella selezionata.")
        return

    # SN di default: prova a dedurlo dal nome cartella (SN_YYYYMMDD_HHMMSS)
    sess_base = os.path.basename(sess_dir)
    default_sn = sess_base.split("_")[0] if "_" in sess_base else "INV"

    made = 0
    for x in xlsx_files:
        x_path = os.path.join(sess_dir, x)
        test_name = os.path.splitext(x)[0]  # es. "curva MPPT DC1"
        # prova a trovare il template omonimo in .\template\
        t_candidate = os.path.join(template_root, f"{test_name}.xlsx")
        if not os.path.isfile(t_candidate):
            print(f"[WARN] Template non trovato per {test_name}: {t_candidate} (skip)")
            continue

        # prova a estrarre il seriale dal workbook (sheet con nome SN) o usa quello di default
        try:
            import pandas as pd
            xl = pd.ExcelFile(x_path, engine="openpyxl")
            serials = [s for s in xl.sheet_names if s and len(s) >= 8 and "_LogErrori" not in s]
            if serials:
                sn = serials[0]
            else:
                sn = default_sn
        except Exception:
            sn = default_sn

        out_html = os.path.join(sess_dir, f"Report_{test_name}_{sn}.html")
        # genera HTML (+ png) con layout nuovo
        try:
            reports_dir = os.path.join("./Reports", os.path.basename(sess_dir))
            os.makedirs(reports_dir, exist_ok=True)
            out_html = os.path.join(reports_dir, f"Report_{test_name}_{sn}.html")
            out_pdf = os.path.splitext(out_html)[0] + ".pdf"
            _, pdf_path, _ = render_mppt_report_html(
                log_xlsx_path=x_path,
                inverter_serials=[sn],
                template_path=t_candidate,
                out_html_path=out_html,
                out_pdf_path=out_pdf,
                logo_path=None, header_path=None, footer_path=None,
                meta={"company": "GID Lab"}
            )
            print("[INFO] Report creato:", out_html)
            made += 1
        except Exception as e:
            print("[ERR] Report fallito per", x_path, "->", e)

    # crea indice sessione con link a tutti i report/test
    try:
        idx_html, idx_pdf = render_session_index(sess_dir, company="GID Lab")
        print("[INFO] Index sessione:", idx_html)
        messagebox.showinfo("Report sessione", f"Creati {made} report.\nIndex: {idx_html}")
        # apri l'index
        try:
            if os.name == "nt":
                os.startfile(idx_html)  # type: ignore
            elif sys.platform.startswith("darwin"):
                subprocess.Popen(["open", idx_html])
            else:
                subprocess.Popen(["xdg-open", idx_html])
        except Exception:
            pass
    except Exception as e:
        messagebox.showwarning("Index", f"Index non creato: {e}")



def on_kill_exit():
    root.destroy()  # Chiude tutto


# Creazione finestra principale
root = tk.Tk()
root.title("Pannello di Controllo")
root.minsize(300, 100)

# global playlist_path_var
# playlist_path_var = tk.StringVar(master=root, value="")

# Lista bottoni con testo, funzione e colore
buttons = [
    ("Controllo DC", on_dc_control, "gold"),
    ("Controllo AC", on_ac_control, "gold"),
    ("Controllo Inverter", on_inverter_control, "gold"),
    ("Lancio Log Strumenti", on_log_start, "lightgreen"),
    #("Lancio Test Automatici", on_test_start, "lightgreen"),
    #("Monitoraggio Realtime", on_realtime_monitor, "lightblue"),
    ("Generazione Report Singolo", on_generate_report, "lightblue"),
    ("Generazione Report Sessione", on_generate_session_report, "lightblue"),
    ("Exit", on_kill_exit, "red")
]

# Inserimento dei bottoni nel pannello
for text, command, color in buttons:
    fg_color = "white" if color == "red" else "black"
    btn = tk.Button(root, text=text, command=command,
                    font=("Arial", 12), bg=color, fg=fg_color, padx=10, pady=10)
    btn.pack(fill="x", padx=30, pady=5)

root.mainloop()
